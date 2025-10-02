import cv2
import numpy as np
import base64
import os
import pandas as pd
from datetime import datetime
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_from_directory, make_response
from functools import wraps
from models import db, Student, Attendance, Faculty, Syllabus, Timetable, SubjectCompletion
from sqlalchemy import func, and_
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import Column, Integer, String, ForeignKey
from sqlalchemy.orm import relationship
import face_recognition  # Add face_recognition import
import time
import json
import zipfile
import tempfile
import glob
import shutil

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this to a secure secret key
# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///face_attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# Constants from Kivy app
DEPARTMENTS = {
    "Microbiology": ["Dr. Chetan Chandrakant Ambasana", "Dr. Jignasha Trikamlal Thumar", 
                     "Dr. Arefa Abdulkhaliq Baakza", "Dr. Sheetal Pramodbhai Pithva", 
                     "Dr. Dharmesh N Adhyaru"],

    "Mathematics": ["Dr. Bhavin Mansukhla Patel", "Dr. Yogita Madhukant Parmar", 
                    "Mrs. Meena Mulchandani", "Mrs. Mamta S Amrutiya"],

    "Physics": ["Dr. Mukesh Vachhani", "Dr. Pooja Uttamprakash Sharma", 
                "Dr. Vibha B Vansola", "Dr. Kashmira P Tank", 
                "Mr. Prahlad Chaudhary", "Dr. Amitkumar Patel", "Mr. Pritesh Khatri"],

    "Botany": ["Dr. Suresh K Patel", "Dr. Nailesh A Patel", 
               "Dr. Pragna Parsottambhai Prajapati", "Dr. Mukesh M Patel", 
               "Dr. Rita Shivlal Chudasama", "Dr. Rohitkumar Patel", 
               "Dr. Binny Karlikar", "Mr. Chandresh R Kharadi"],

    "Zoology": ["Dr. Chetana V Shah", "Dr. Manishkumar D Visavadia", "Dr. Heena Prajapati"],

    "Chemistry": ["Dr. Shaileshkumar Prajapati", "Dr. K.B Patel", "Dr. K.S Nimavat", 
                  "Khachar Natubhai Bhurabhai", "Dr. Harshad P Lakum", 
                  "Dr. Chetan B Sangani", "Miss Nisha Rameshchand Sharma", "Dr. Mamta T Singh"],

    "Statistics": ["Dr. Samir Pandya", "Dr. Dharak Patel"]
}

# Group to Subject mappings
GROUP_SUBJECTS = {
    'PCB': {'major': 'Physics', 'minor': 'Chemistry', 'multi': 'Botany'},
    'PCM': {'major': 'Physics', 'minor': 'Chemistry', 'multi': 'Mathematics'},
    'PMC': {'major': 'Physics', 'minor': 'Mathematics', 'multi': 'Chemistry'},
    'PMS': {'major': 'Physics', 'minor': 'Mathematics', 'multi': 'Statistics'},
    'BCP': {'major': 'Botany', 'minor': 'Chemistry', 'multi': 'Physics'},
    'BCZ': {'major': 'Botany', 'minor': 'Chemistry', 'multi': 'Zoology'},
    'BZC': {'major': 'Botany', 'minor': 'Zoology', 'multi': 'Chemistry'},
    'BZMi': {'major': 'Botany', 'minor': 'Zoology', 'multi': 'Microbiology'},
    'BZM': {'major': 'Botany', 'minor': 'Zoology', 'multi': 'Microbiology'},
    'CBP': {'major': 'Chemistry', 'minor': 'Botany', 'multi': 'Physics'},
    'CBZ': {'major': 'Chemistry', 'minor': 'Botany', 'multi': 'Zoology'},
    'CPM': {'major': 'Chemistry', 'minor': 'Physics', 'multi': 'Mathematics'},
    'MiZB': {'major': 'Microbiology', 'minor': 'Zoology', 'multi': 'Botany'},
    'MZB': {'major': 'Microbiology', 'minor': 'Zoology', 'multi': 'Botany'},
    'MPC': {'major': 'Mathematics', 'minor': 'Physics', 'multi': 'Chemistry'},
    'MPS': {'major': 'Mathematics', 'minor': 'Physics', 'multi': 'Statistics'},
    'SMP': {'major': 'Statistics', 'minor': 'Mathematics', 'multi': 'Physics'},
    'ZBC': {'major': 'Zoology', 'minor': 'Botany', 'multi': 'Chemistry'},
    'ZBMi': {'major': 'Zoology', 'minor': 'Botany', 'multi': 'Microbiology'}
}

SUBJECT_CODES = {
    "Microbiology": ["MB101", "MB102", "MB103"],
    "Mathematics": ["M101", "M102", "M103"],
    "Physics": ["P101", "P102", "P103"],
    "Botany": ["B101", "B102", "B103"],
    "Zoology": ["Z101", "Z102", "Z103"],
    "Chemistry": ["C101", "C102", "C103"],
    "Statistics": ["S101", "S102", "S103"]
}

UNITS = ["Unit 1", "Unit 2", "Unit 3", "Unit 4"]
SEMESTERS = ['Semester 1', 'Semester 2', 'Semester 3', 'Semester 4', 'Semester 5', 'Semester 6']

# Store registered faces and attendance
registered_faces = {}
attendance_records = []

# Faculty credentials
FACULTY_CREDENTIALS = {
    'Microbiology': {
        'CCA001': {'name': 'Dr. Chetan Chandrakant Ambasana', 'password': 'cca@123'},
        'JTT001': {'name': 'Dr. Jignasha Trikamlal Thumar', 'password': 'jtt@123'},
        'AAB001': {'name': 'Dr. Arefa Abdulkhaliq Baakza', 'password': 'aab@123'},
        'SPP001': {'name': 'Dr. Sheetal Pramodbhai Pithva', 'password': 'spp@123'},
        'DNA001': {'name': 'Dr. Dharmesh N Adhyaru', 'password': 'dna@123'}
    },
    'Mathematics': {
        'BMP001': {'name': 'Dr. Bhavin Mansukhla Patel', 'password': 'bmp@123'},
        'YMP001': {'name': 'Dr. Yogita Madhukant Parmar', 'password': 'ymp@123'},
        'MM001': {'name': 'Mrs. Meena Mulchandani', 'password': 'mm@123'},
        'MSA001': {'name': 'Mrs. Mamta S Amrutiya', 'password': 'msa@123'}
    },
    'Physics': {
        'MV001': {'name': 'Dr. Mukesh Vachhani', 'password': 'mv@123'},
        'PUS001': {'name': 'Dr. Pooja Uttamprakash Sharma', 'password': 'pus@123'},
        'VBV001': {'name': 'Dr. Vibha B Vansola', 'password': 'vbv@123'},
        'KPT001': {'name': 'Dr. Kashmira P Tank', 'password': 'kpt@123'},
        'PC001': {'name': 'Mr. Prahlad Chaudhary', 'password': 'pc@123'},
        'AP001': {'name': 'Dr. Amitkumar Patel', 'password': 'ap@123'},
        'PK001': {'name': 'Mr. Pritesh Khatri', 'password': 'pk@123'}
    },
    'Botany': {
        'SKP001': {'name': 'Dr. Suresh K Patel', 'password': 'skp@123'},
        'NAP001': {'name': 'Dr. Nailesh A Patel', 'password': 'nap@123'},
        'PPP001': {'name': 'Dr. Pragna Parsottambhai Prajapati', 'password': 'ppp@123'},
        'MMP001': {'name': 'Dr. Mukesh M Patel', 'password': 'mmp@123'},
        'RSC001': {'name': 'Dr. Rita Shivlal Chudasama', 'password': 'rsc@123'},
        'RP001': {'name': 'Dr. Rohitkumar Patel', 'password': 'rp@123'},
        'BK001': {'name': 'Dr. Binny Karlikar', 'password': 'bk@123'},
        'CRK001': {'name': 'Mr. Chandresh R Kharadi', 'password': 'crk@123'}
    },
    'Zoology': {
        'CVS001': {'name': 'Dr. Chetana V Shah', 'password': 'cvs@123'},
        'MDV001': {'name': 'Dr. Manishkumar D Visavadia', 'password': 'mdv@123'},
        'HP001': {'name': 'Dr. Heena Prajapati', 'password': 'hp@123'}
    },
    'Chemistry': {
        'SP001': {'name': 'Dr. Shaileshkumar Prajapati', 'password': 'sp@123'},
        'KBP001': {'name': 'Dr. K.B Patel', 'password': 'kbp@123'},
        'KSN001': {'name': 'Dr. K.S Nimavat', 'password': 'ksn@123'},
        'KNB001': {'name': 'Khachar Natubhai Bhurabhai', 'password': 'knb@123'},
        'HPL001': {'name': 'Dr. Harshad P Lakum', 'password': 'hpl@123'},
        'CBS001': {'name': 'Dr. Chetan B Sangani', 'password': 'cbs@123'},
        'NRS001': {'name': 'Miss Nisha Rameshchand Sharma', 'password': 'nrs@123'},
        'MTS001': {'name': 'Dr. Mamta T Singh', 'password': 'mts@123'}
    },
    'Statistics': {
        'SP002': {'name': 'Dr. Samir Pandya', 'password': 'sp@123'},
        'DP001': {'name': 'Dr. Dharak Patel', 'password': 'dp@123'}
    }
}

# Admin credentials (you should move these to a secure configuration file in production)
ADMIN_CREDENTIALS = {
    'username': 'admin',
    'password': 'admin123'  # In production, use a strong hashed password
}

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'faculty_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin session management
def admin_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login_page'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        faculty = Faculty.query.filter_by(username=username).first()
        
        if faculty and faculty.check_password(password):
            session['faculty_id'] = faculty.id
            session['faculty_name'] = faculty.name
            session['department'] = faculty.department
            session['is_hod'] = faculty.is_hod  # Add HOD status to session
            return redirect(url_for('index'))
            
        return render_template('login.html', error='Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    is_hod = session.get('is_hod', False)
    department = session.get('department')
    return render_template('index.html', 
                         faculty_name=session.get('faculty_name'),
                         department=department,
                         is_hod=is_hod)

@app.route('/get_faculty/<department>')
def get_faculty(department):
    faculty_list = DEPARTMENTS.get(department, [])
    return jsonify(faculty_list)

@app.route('/get_department_subjects/<department>')
def get_department_subjects(department):
    subject_list = SUBJECT_CODES.get(department, [])
    return jsonify(subject_list)

@app.route('/register', methods=['POST'])
@admin_login_required
def register():
    try:
        data = request.json
        if not data or 'images' not in data or len(data['images']) != 3:
            return jsonify({'success': False, 'message': 'Three photos are required for registration'})

        required_fields = ['name', 'enrollment', 'semester', 'group', 'major_subject', 'minor_subject', 'multi_subject']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Missing required field: {field}'})

        # Process all three images and store their encodings
        face_encodings = []
        photos = []
        
        for img_data in data['images']:
            img = process_base64_image(img_data)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_recognition.face_locations(gray)

            if len(faces) == 0:
                return jsonify({'success': False, 'message': 'No face detected in one or more photos'})
            
            if len(faces) > 1:
                    return jsonify({'success': False, 'message': 'Multiple faces detected in one photo. Please capture one face at a time'})
            
            # Convert image to binary for storage
            _, img_encoded = cv2.imencode('.jpg', img)
            photos.append(img_encoded.tobytes())
            face_encodings.append(faces[0])

        # Create new student record
        new_student = Student(
            name=data['name'],
            enrollment_number=data['enrollment'],
            semester=data['semester'],
            group=data['group'],
            major_subject=data['major_subject'],
            minor_subject=data['minor_subject'],
            multi_subject=data['multi_subject'],
            photo=photos[0],  # Store the first photo as main photo
            face_encoding=face_encodings  # Store all face encodings
        )

        # Add and commit to database
        db.session.add(new_student)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Successfully registered {new_student.name} with 3 photos'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Registration error: {str(e)}")
        return jsonify({'success': False, 'message': f'Error during registration: {str(e)}'})

@app.route('/mark-attendance', methods=['POST'])
@login_required
def mark_attendance():
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({'success': False, 'message': 'No image data received'})

        # Process the base64 image from the browser
        try:
            # Remove the data URL prefix if present
            image_data = data['image']
            if ',' in image_data:
                image_data = image_data.split(',')[1]
            
            # Decode base64 to image
            img_bytes = base64.b64decode(image_data)
            nparr = np.frombuffer(img_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if img is None:
                return jsonify({'success': False, 'message': 'Invalid image data'})
            
            # Convert BGR to RGB
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            print("Successfully decoded and converted image")
            
        except Exception as e:
            print(f"Error processing image: {str(e)}")
            return jsonify({'success': False, 'message': 'Error processing image'})

        # Get students from the selected semester only
        students = Student.query.filter_by(semester=data['semester']).all()
        if not students:
            return jsonify({
                'success': False,
                'message': f'No students found in {data["semester"]}'
            })
        print(f"Found {len(students)} students in semester {data['semester']}")

        recognized_students = set()  # Use set to avoid duplicate attendance

        # For each detected face, compare with registered students
        face_locations = face_recognition.face_locations(img_rgb)
        print(f"Detected {len(face_locations)} faces in the image")
        
        if not face_locations:
            return jsonify({'success': False, 'message': 'No faces detected in the image'})
            
        face_encodings = face_recognition.face_encodings(img_rgb, face_locations)
        print(f"Generated {len(face_encodings)} face encodings")

        for face_encoding in face_encodings:
            for student in students:
                if student.face_encoding is None:
                    print(f"Student {student.name} has no face encoding")
                    continue
                
                # Compare face encodings with stored encodings
                matches = []
                for stored_encoding in student.face_encoding:
                    # Compare faces with a stricter tolerance (default is 0.6)
                    match = face_recognition.compare_faces([stored_encoding], face_encoding, tolerance=0.5)[0]
                    matches.append(match)
                
                # If any of the stored photos match
                if any(matches) and student.id not in recognized_students:
                    print(f"Matched face with student {student.name}")
                    attendance = Attendance(
                        student_id=student.id,
                        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
                        time=datetime.strptime(data['time'], '%H:%M').time(),
                        department=data['department'],
                        faculty=data['faculty'],
                        subject=data['subject'],
                        unit=data['unit']
                    )
                    db.session.add(attendance)
                    recognized_students.add(student.id)

        if recognized_students:
            db.session.commit()
            recognized_list = [
                {'enrollment_number': student.enrollment_number, 'name': student.name}
                for student in students if student.id in recognized_students
            ]
            print(f"Successfully recorded attendance for {len(recognized_students)} students")
            return jsonify({
                'success': True,
                'message': f'Recorded attendance for {len(recognized_students)} students from {data["semester"]}',
                'recognized_students': recognized_list
            })
        else:
            print("No students were recognized")
            return jsonify({
                'success': False,
                'message': f'No registered students from {data["semester"]} were recognized'
            })

    except Exception as e:
        db.session.rollback()
        print(f"Attendance marking error: {str(e)}")
        return jsonify({'success': False, 'message': f'Error marking attendance: {str(e)}'})

@app.route('/export-attendance', methods=['POST'])
@login_required
def export_attendance():
    try:
        department = request.json.get('department', '')
        if not department:
            return jsonify({
                'success': False,
                'message': 'Department is required'
            })

        # Create exports directory if it doesn't exist
        current_dir = os.path.dirname(os.path.abspath(__file__))
        exports_dir = os.path.join(current_dir, 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Define fixed filename for each department in exports directory
        try:
            filename = os.path.join(exports_dir, f"attendance_{department}.xlsx")
            print(f"Will save Excel file to: {filename}")
        except Exception as e:
            print(f"Error creating file path: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to create file path: {str(e)}'
            })

        # Query attendance records
        try:
            attendance_data = (
                db.session.query(
                    Student.name,
                    Student.enrollment_number,
                    Student.semester,
                    Attendance.date,
                    Attendance.time,
                    Attendance.department,
                    Attendance.faculty,
                    Attendance.subject,
                    Attendance.unit
                )
                .select_from(Attendance)
                .join(Student, Student.id == Attendance.student_id)
                .filter(Attendance.department == department)
                .order_by(
                    Attendance.date.desc(),
                    Attendance.time.desc(),
                    Student.enrollment_number
                )
                .all()
            )
            print(f"Found {len(attendance_data)} attendance records")
        except Exception as e:
            print(f"Database query error: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to query attendance data: {str(e)}'
            })

        if not attendance_data:
            print("No attendance records found")
            return jsonify({
                'success': False,
                'message': 'No attendance records found for the selected department'
            })

        # Convert to DataFrame
        try:
            records = []
            for record in attendance_data:
                record_dict = {
                    'Enrollment Number': record.enrollment_number,
                    'Student Name': record.name,
                    'Semester': record.semester,
                    'Date': record.date.strftime('%Y-%m-%d'),
                    'Time': record.time.strftime('%H:%M'),
                    'Faculty': record.faculty,
                    'Subject': record.subject,
                    'Unit': record.unit
                }
                records.append(record_dict)
                print(f"Processing record: {record_dict}")

            df = pd.DataFrame(records)
            df = df.sort_values(['Date', 'Time', 'Enrollment Number'], 
                              ascending=[False, False, True])

            column_order = [
                'Enrollment Number',
                'Student Name',
                'Semester',
                'Date',
                'Time',
                'Faculty',
                'Subject',
                'Unit'
            ]
            df = df[column_order]
            print(f"Created DataFrame with shape: {df.shape}")
        except Exception as e:
            print(f"Error creating DataFrame: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to process attendance data: {str(e)}'
            })

        # Save Excel file
        try:
            # First try to remove the file if it exists
            if os.path.exists(filename):
                os.remove(filename)
                print(f"Removed existing file: {filename}")

            # Save the new file
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Attendance Records')
                workbook = writer.book
                worksheet = writer.sheets['Attendance Records']

                # Format header
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Format cells
                for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    fill_color = 'F0F8FF' if (idx - 2) % 2 == 0 else 'FFFFFF'
                    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        cell.fill = row_fill

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

            print(f"Successfully saved Excel file to: {filename}")
            return jsonify({
                'success': True,
                'message': f'Attendance exported to exports/attendance_{department}.xlsx'
            })

        except Exception as e:
            print(f"Error saving Excel file: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to save Excel file: {str(e)}'
            })

    except Exception as e:
        print(f"General export error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to export attendance: {str(e)}'
        })

def process_base64_image(base64_string):
    try:
        # Ensure base64 string is valid
        if not base64_string or not isinstance(base64_string, str):
            print("Error: Base64 string is empty or not a string")
            return None

        # Check if base64 string contains metadata (e.g., "data:image/png;base64,")
        if "," in base64_string:
            base64_string = base64_string.split(",")[1]  # Remove metadata

        # Decode base64 to bytes
        image_data = base64.b64decode(base64_string)
        np_arr = np.frombuffer(image_data, np.uint8)
        
        # Read image from buffer
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)  # Read as color image

        if img is None:
            print("Error: Decoded image is None (possibly corrupt)")
            return None

        # Convert BGR to RGB
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        return img_rgb

    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return None

@app.route('/register-student')
@admin_login_required
def register_student():
    return render_template('register.html',
                         departments=DEPARTMENTS,
                         subject_codes=SUBJECT_CODES,
                         units=UNITS,
                         semesters=SEMESTERS)

@app.route('/mark-attendance-page')
@login_required
def mark_attendance_page():
    return render_template('attendance.html',
                         departments=DEPARTMENTS,
                         semesters=SEMESTERS,
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/export-attendance-page')
@login_required
def export_attendance_page():
    return render_template('export.html',
                         departments=DEPARTMENTS,
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/get-semester-students/<semester>')
@login_required
def get_semester_students(semester):
    try:
        students = Student.query.filter_by(semester=semester).all()
        student_list = [{
            'id': student.id,
            'name': student.name,
            'enrollment_number': student.enrollment_number
        } for student in students]
        
        return jsonify({
            'success': True,
            'students': student_list
        })
    except Exception as e:
        print(f"Error fetching students: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching students: {str(e)}'
        })

@app.route('/mark-student-present', methods=['POST'])
@login_required
def mark_student_present():
    try:
        data = request.json
        if not data or 'student_id' not in data:
            return jsonify({
                'success': False,
                'message': 'Student ID is required'
            })

        student = Student.query.get(data['student_id'])
        if not student:
            return jsonify({
                'success': False,
                'message': 'Student not found'
            })

        # Create attendance record
        attendance = Attendance(
            student_id=student.id,
            date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
            time=datetime.strptime(data['time'], '%H:%M').time(),
            department=data['department'],
            faculty=data['faculty'],
            subject=data['subject'],
            unit=data['unit']
        )
        db.session.add(attendance)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Marked {student.name} as present',
            'name': student.name,
            'enrollment_number': student.enrollment_number
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error marking student present: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error marking student present: {str(e)}'
        })

@app.route('/get-attendance-history', methods=['POST'])
@login_required
def get_attendance_history():
    try:
        data = request.json
        query = db.session.query(
            Attendance.date,
            Attendance.time,
            Student.semester,
            Attendance.subject,
            Attendance.unit,
            func.count(Attendance.student_id).label('present_count')
        ).join(
            Student,
            Student.id == Attendance.student_id
        ).group_by(
            Attendance.date,
            Attendance.time,
            Student.semester,
            Attendance.subject,
            Attendance.unit
        ).filter(
            Attendance.faculty == data['faculty'],
            Attendance.department == data['department']
        )

        # Apply filters if provided
        if data.get('startDate') and data.get('endDate'):
            start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['endDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date.between(start_date, end_date))
        elif data.get('startDate'):
            start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date >= start_date)
        elif data.get('endDate'):
            end_date = datetime.strptime(data['endDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date <= end_date)

        if data.get('semester'):
            query = query.filter(Student.semester == data['semester'])
        if data.get('subject'):
            query = query.filter(Attendance.subject == data['subject'])

        records = query.order_by(Attendance.date.desc(), Attendance.time.desc()).all()

        # Calculate absent count for each record
        formatted_records = []
        for record in records:
            # Get total students in the semester
            total_students = Student.query.filter_by(semester=record.semester).count()
            
            record_dict = {
                'date': record.date.strftime('%Y-%m-%d'),
                'time': record.time.strftime('%H:%M'),
                'semester': record.semester,
                'subject': record.subject,
                'unit': record.unit,
                'present_count': record.present_count,
                'absent_count': total_students - record.present_count
            }
            formatted_records.append(record_dict)

        return jsonify({
            'success': True,
            'records': formatted_records
        })

    except Exception as e:
        print(f"Error fetching attendance history: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching attendance history: {str(e)}'
        })

@app.route('/attendance-details')
@login_required
def attendance_details():
    try:
        date = request.args.get('date')
        semester = request.args.get('semester')
        subject = request.args.get('subject')
        
        if not all([date, semester, subject]):
            print("Missing required parameters")
            return redirect(url_for('attendance_history_page'))

        # Get attendance records for the specific date, semester, and subject
        attendance_records = db.session.query(
            Student.name,
            Student.enrollment_number,
            Attendance.time,
            Attendance.unit
        ).join(
            Attendance,
            Student.id == Attendance.student_id
        ).filter(
            Attendance.date == datetime.strptime(date, '%Y-%m-%d').date(),
            Student.semester == semester,
            Attendance.subject == subject,
            Attendance.faculty == session.get('faculty_name'),
            Attendance.department == session.get('department')
        ).all()

        # Get all students in the semester
        all_students = Student.query.filter_by(semester=semester).all()
        
        # Create a set of present student enrollment numbers
        present_enrollments = {record.enrollment_number for record in attendance_records}
        
        # Create present and absent lists
        present_students = attendance_records
        absent_students = [student for student in all_students 
                         if student.enrollment_number not in present_enrollments]

        # Get unique units from attendance records
        units = sorted(list(set(record.unit for record in attendance_records)))

        return render_template(
            'attendance_details.html',
            date=date,
            semester=semester,
            subject=subject,
            present_students=present_students,
            absent_students=absent_students,
            units=units,
            faculty_name=session.get('faculty_name'),
            department=session.get('department')
        )

    except Exception as e:
        print(f"Error in attendance_details: {str(e)}")
        return redirect(url_for('attendance_history_page'))

@app.route('/exports/<path:filename>')
@login_required
def serve_export(filename):
    try:
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        return send_from_directory(exports_dir, filename, as_attachment=True)
    except Exception as e:
        print(f"Error serving export: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error serving export: {str(e)}'
        })

@app.route('/export-specific-attendance', methods=['POST'])
@login_required
def export_specific_attendance():
    try:
        data = request.json
        date = datetime.strptime(data['date'], '%Y-%m-%d').date()

        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Query attendance records
        attendance_records = db.session.query(
            Student.enrollment_number,
            Student.name,
            Attendance.time,
            Attendance.unit
        ).join(
            Attendance,
            Student.id == Attendance.student_id
        ).filter(
            Attendance.date == date,
            Student.semester == data['semester'],
            Attendance.subject == data['subject'],
            Attendance.faculty == data['faculty'],
            Attendance.department == data['department']
        ).all()

        # Get all students in the semester
        all_students = Student.query.filter_by(semester=data['semester']).all()
        
        # Create records list
        records = []
        present_students = {record.enrollment_number: record for record in attendance_records}
        
        for student in all_students:
            record = present_students.get(student.enrollment_number)
            records.append({
                'Enrollment Number': student.enrollment_number,
                'Student Name': student.name,
                'Status': 'Present' if record else 'Absent',
                'Time': record.time.strftime('%H:%M') if record else '-',
                'Unit': record.unit if record else '-'
            })

        # Create DataFrame and save to Excel
        df = pd.DataFrame(records)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"attendance_{data['department']}_{data['semester']}_{data['subject']}_{date}_{timestamp}.xlsx"
        filepath = os.path.join(exports_dir, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance')
            workbook = writer.book
            worksheet = writer.sheets['Attendance']

            # Format header
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format status column
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, min_col=3, max_col=3), start=2):
                cell = row[0]
                if cell.value == 'Present':
                    cell.font = Font(color='008000')  # Green
                else:
                    cell.font = Font(color='FF0000')  # Red

            # Adjust column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

        return jsonify({
            'success': True,
            'message': 'Attendance exported successfully',
            'file_path': filename
        })

    except Exception as e:
        print(f"Error exporting specific attendance: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error exporting attendance: {str(e)}'
        })

@app.route('/export-all-attendance', methods=['POST'])
@login_required
def export_all_attendance():
    try:
        data = request.json
        
        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Build query with filters
        query = db.session.query(
            Student.name,
            Student.enrollment_number,
            Student.semester,
            Attendance.date,
            Attendance.time,
            Attendance.subject,
            Attendance.unit
        ).join(
            Attendance,
            Student.id == Attendance.student_id
        ).filter(
            Attendance.faculty == data['faculty'],
            Attendance.department == data['department']
        )

        # Apply filters if provided
        if data.get('startDate') and data.get('endDate'):
            start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['endDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date.between(start_date, end_date))
        if data.get('semester'):
            query = query.filter(Attendance.semester == data['semester'])
        if data.get('subject'):
            query = query.filter(Attendance.subject == data['subject'])

        # Execute query
        attendance_records = query.order_by(
            Attendance.date.desc(),
            Attendance.time.desc(),
            Student.enrollment_number
        ).all()

        if not attendance_records:
            return jsonify({
                'success': False,
                'message': 'No attendance records found for the selected criteria'
            })

        # Create DataFrame
        records = []
        for record in attendance_records:
            records.append({
                'Date': record.date.strftime('%Y-%m-%d'),
                'Time': record.time.strftime('%H:%M'),
                'Enrollment Number': record.enrollment_number,
                'Student Name': record.name,
                'Semester': record.semester,
                'Subject': record.subject,
                'Unit': record.unit
            })

        df = pd.DataFrame(records)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"attendance_report_{data['department']}_{timestamp}.xlsx"
        filepath = os.path.join(exports_dir, filename)

        # Save to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance Records')
            workbook = writer.book
            worksheet = writer.sheets['Attendance Records']

            # Format header
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format cells
            for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                fill_color = 'F0F8FF' if (idx - 2) % 2 == 0 else 'FFFFFF'
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = row_fill

            # Adjust column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

        return jsonify({
            'success': True,
            'message': 'Attendance report generated successfully',
            'file_path': filename
        })

    except Exception as e:
        print(f"Error exporting attendance report: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error generating report: {str(e)}'
        })

@app.route('/admin')
def admin_login_page():
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    try:
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            
            if not username or not password:
                return render_template('admin_login.html', error='Please provide both username and password')
            
            if username == ADMIN_CREDENTIALS['username'] and password == ADMIN_CREDENTIALS['password']:
                session['admin_logged_in'] = True
                return redirect(url_for('admin_dashboard'))
            
            return render_template('admin_login.html', error='Invalid credentials')
        
        # If it's a GET request, redirect to the login page
        return redirect(url_for('admin_login_page'))
    except Exception as e:
        print(f"Admin login error: {str(e)}")
        return render_template('admin_login.html', error='An error occurred. Please try again.')

@app.route('/admin/dashboard')
@admin_login_required
def admin_dashboard():
    try:
        # Get total counts
        total_faculty = Faculty.query.count()
        total_students = Student.query.count()
        
        # Get department-wise counts
        department_stats = {}
        for dept in DEPARTMENTS.keys():
            faculty_count = Faculty.query.filter_by(department=dept).count()
            student_count = Student.query.filter(
                Student.major_subject.like(f"%{dept}%")
            ).count()
            department_stats[dept] = {
                'faculty': faculty_count,
                'students': student_count
            }
        
        return render_template('admin_dashboard.html',
                             total_faculty=total_faculty,
                             total_students=total_students,
                             department_stats=department_stats,
                             semesters=SEMESTERS)
    except Exception as e:
        print(f"Admin dashboard error: {str(e)}")
        return redirect(url_for('admin_login_page'))

@app.route('/admin/logout')
def admin_logout():
    try:
        session.pop('admin_logged_in', None)
        return redirect(url_for('admin_login_page'))
    except Exception as e:
        print(f"Admin logout error: {str(e)}")
        return redirect(url_for('admin_login_page'))

@app.route('/admin/manage-faculty')
@admin_login_required
def manage_faculty():
    try:
        faculty_members = Faculty.query.all()
        departments = DEPARTMENTS.keys()
        return render_template('manage_faculty.html', 
                             faculty_members=faculty_members,
                             departments=departments)
    except Exception as e:
        print(f"Error fetching faculty members: {str(e)}")
        return render_template('manage_faculty.html', 
                             faculty_members=[], 
                             departments=DEPARTMENTS.keys(),
                             error="Error fetching faculty members")

@app.route('/admin/toggle-hod/<int:faculty_id>', methods=['POST'])
@admin_login_required
def toggle_hod(faculty_id):
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        department = faculty.department

        # If making this faculty HOD, remove HOD status from other faculty in same department
        if not faculty.is_hod:
            current_hod = Faculty.query.filter_by(department=department, is_hod=True).first()
            if current_hod:
                current_hod.is_hod = False
                db.session.add(current_hod)

        # Toggle HOD status for selected faculty
        faculty.is_hod = not faculty.is_hod
        db.session.add(faculty)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{"Added as" if faculty.is_hod else "Removed as"} HOD of {department} department',
            'is_hod': faculty.is_hod
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error toggling HOD status: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error updating HOD status: {str(e)}'
        })

@app.route('/admin/get-department-hod/<department>')
@admin_login_required
def get_department_hod(department):
    try:
        hod = Faculty.query.filter_by(department=department, is_hod=True).first()
        return jsonify({
            'success': True,
            'hod': {
                'id': hod.id,
                'name': hod.name
            } if hod else None
        })
    except Exception as e:
        print(f"Error fetching HOD: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching HOD: {str(e)}'
        })

@app.route('/admin/add-faculty', methods=['POST'])
@admin_login_required
def add_faculty():
    try:
        name = request.form.get('name')
        department = request.form.get('department')
        email = request.form.get('email')
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Check if username already exists
        if Faculty.query.filter_by(username=username).first():
            return jsonify({'success': False, 'message': 'Username already exists'})
        
        # Create new faculty member
        faculty = Faculty(
            name=name,
            department=department,
            email=email,
            username=username
        )
        faculty.set_password(password)
        
        db.session.add(faculty)
        db.session.commit()
        
        return redirect(url_for('manage_faculty'))
    except Exception as e:
        db.session.rollback()
        print(f"Error adding faculty: {str(e)}")
        return jsonify({'success': False, 'message': f'Error adding faculty: {str(e)}'})

@app.route('/admin/delete-faculty/<int:faculty_id>', methods=['POST'])
@admin_login_required
def delete_faculty(faculty_id):
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        db.session.delete(faculty)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting faculty: {str(e)}")
        return jsonify({'success': False, 'message': f'Error deleting faculty: {str(e)}'})

@app.route('/admin/manage-syllabus')
@admin_login_required
def manage_syllabus():
    try:
        # Get all syllabus entries
        syllabus_entries = Syllabus.query.all()
        
        # Get unique departments from the constants
        departments = list(DEPARTMENTS.keys())
        
        return render_template('manage_syllabus.html',
                             departments=departments,
                             semesters=SEMESTERS,
                             syllabus_entries=syllabus_entries)
                             
    except Exception as e:
        print(f"Error in manage_syllabus route: {str(e)}")
        return render_template('manage_syllabus.html', 
                             departments=list(DEPARTMENTS.keys()),
                             semesters=SEMESTERS,
                             syllabus_entries=[],
                             error="Error fetching syllabus entries")

@app.route('/admin/add-syllabus', methods=['POST'])
@admin_login_required
def add_syllabus():
    try:
        department = request.form.get('department')
        semester = request.form.get('semester')
        subject = request.form.get('subject')
        units = request.form.getlist('units[]')  # Get multiple units
        
        # Join units with comma
        units_str = ','.join(units)
        
        # Check if entry already exists
        existing = Syllabus.query.filter_by(
            department=department,
            semester=semester,
            subject=subject
        ).first()
        
        if existing:
            # Update existing entry
            existing.units = units_str
            db.session.commit()
            return jsonify({'success': True, 'message': 'Syllabus updated successfully'})
        
        # Create new syllabus entry
        syllabus = Syllabus(
            department=department,
            semester=semester,
            subject=subject,
            units=units_str
        )
        
        db.session.add(syllabus)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Syllabus added successfully'})
    except Exception as e:
        db.session.rollback()
        print(f"Error adding syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error adding syllabus: {str(e)}'})

@app.route('/admin/delete-syllabus/<int:syllabus_id>', methods=['POST'])
@admin_login_required
def delete_syllabus(syllabus_id):
    try:
        syllabus = Syllabus.query.get_or_404(syllabus_id)
        db.session.delete(syllabus)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Syllabus deleted successfully'})
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error deleting syllabus: {str(e)}'})

@app.route('/admin/get-syllabus/<department>/<semester>')
@admin_login_required
def get_syllabus(department, semester):
    try:
        syllabus_entries = Syllabus.query.filter_by(
            department=department,
            semester=semester
        ).all()
        
        entries = [{
            'id': entry.id,
            'subject': entry.subject,
            'units': entry.units.split(',')
        } for entry in syllabus_entries]
        
        return jsonify({'success': True, 'entries': entries})
    except Exception as e:
        print(f"Error fetching syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching syllabus: {str(e)}'})

@app.route('/admin/update-syllabus/<int:syllabus_id>', methods=['POST'])
@admin_login_required
def update_syllabus(syllabus_id):
    try:
        data = request.json
        syllabus = Syllabus.query.get_or_404(syllabus_id)
        
        # Update subject if provided
        if 'subject' in data:
            # Check if the new subject already exists for this department and semester
            existing = Syllabus.query.filter_by(
                department=syllabus.department,
                semester=syllabus.semester,
                subject=data['subject']
            ).first()
            
            if existing and existing.id != syllabus_id:
                return jsonify({
                    'success': False,
                    'message': 'This subject already exists for the selected department and semester'
                })
            
            syllabus.subject = data['subject']
        
        # Update units
        units = data.get('units', [])
        syllabus.units = ','.join(units)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Syllabus updated successfully'})
    except Exception as e:
        db.session.rollback()
        print(f"Error updating syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error updating syllabus: {str(e)}'})

@app.route('/admin/update-semester', methods=['POST'])
@admin_login_required
def update_semester():
    try:
        data = request.json
        current_semester = data.get('current_semester')
        
        if not current_semester:
            return jsonify({
                'success': False,
                'message': 'Current semester is required'
            })

        # Get the index of current semester
        try:
            current_index = SEMESTERS.index(current_semester)
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Invalid semester'
            })

        # Check if this is the last semester
        if current_index >= len(SEMESTERS) - 1:
            return jsonify({
                'success': False,
                'message': f'Students in {current_semester} cannot be promoted further'
            })

        # Get the next semester
        next_semester = SEMESTERS[current_index + 1]

        # Update students' semester
        students = Student.query.filter_by(semester=current_semester).all()
        
        # Print for debugging
        print(f"Found {len(students)} students in {current_semester}")
        
        if not students:
            print(f"No students found in semester: {current_semester}")
            return jsonify({
                'success': False,
                'message': f'No students found in {current_semester}'
            })

        # Update each student's semester
        updated_count = 0
        for student in students:
            student.semester = next_semester
            updated_count += 1
            print(f"Updated student: {student.name} from {current_semester} to {next_semester}")

        db.session.commit()
        print(f"Successfully updated {updated_count} students")

        return jsonify({
            'success': True,
            'message': f'Successfully promoted {updated_count} students from {current_semester} to {next_semester}'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error updating semesters: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error updating semesters: {str(e)}'
        })

# Create database tables and initialize faculty data
def init_faculty_data():
    try:
        # Check if faculty data already exists
        if Faculty.query.count() == 0:
            print("Initializing faculty data...")
            # Add faculty members from FACULTY_CREDENTIALS
            for department, faculty_dict in FACULTY_CREDENTIALS.items():
                for faculty_id, details in faculty_dict.items():
                    # Create faculty member
                    faculty = Faculty(
                        name=details['name'],
                        department=department,
                        email=f"{faculty_id.lower()}@college.edu",  # Generate email from faculty ID
                        username=faculty_id
                    )
                    faculty.set_password(details['password'])
                    db.session.add(faculty)
            
            db.session.commit()
            print("Faculty data initialized successfully!")
    except Exception as e:
        db.session.rollback()
        print(f"Error initializing faculty data: {str(e)}")

def init_syllabus_data():
    try:
        # Check if syllabus data already exists
        if Syllabus.query.count() == 0:
            print("Initializing syllabus data...")
            
            # Define semester-wise subjects for each department
            department_subjects = {
                "Physics": {
                    "Semester 1": ["Mechanics", "Properties of Matter", "Physics Lab-1"],
                    "Semester 2": ["Waves and Oscillations", "Thermodynamics", "Physics Lab-2"],
                    "Semester 3": ["Optics", "Electricity and Magnetism", "Physics Lab-3"],
                    "Semester 4": ["Modern Physics", "Electronics", "Physics Lab-4"],
                    "Semester 5": ["Quantum Mechanics", "Nuclear Physics", "Physics Lab-5"],
                    "Semester 6": ["Solid State Physics", "Atomic Physics", "Physics Lab-6"]
                },
                "Chemistry": {
                    "Semester 1": ["Inorganic Chemistry", "Physical Chemistry", "Chemistry Lab-1"],
                    "Semester 2": ["Organic Chemistry", "Analytical Chemistry", "Chemistry Lab-2"],
                    "Semester 3": ["Coordination Chemistry", "Chemical Kinetics", "Chemistry Lab-3"],
                    "Semester 4": ["Organic Synthesis", "Spectroscopy", "Chemistry Lab-4"],
                    "Semester 5": ["Advanced Inorganic", "Quantum Chemistry", "Chemistry Lab-5"],
                    "Semester 6": ["Biochemistry", "Industrial Chemistry", "Chemistry Lab-6"]
                },
                "Mathematics": {
                    "Semester 1": ["Calculus", "Algebra", "Discrete Mathematics"],
                    "Semester 2": ["Real Analysis", "Linear Algebra", "Number Theory"],
                    "Semester 3": ["Complex Analysis", "Differential Equations", "Numerical Methods"],
                    "Semester 4": ["Abstract Algebra", "Topology", "Operations Research"],
                    "Semester 5": ["Graph Theory", "Probability Theory", "Mathematical Modeling"],
                    "Semester 6": ["Functional Analysis", "Statistics", "Advanced Calculus"]
                },
                "Botany": {
                    "Semester 1": ["Plant Diversity", "Cell Biology", "Botany Lab-1"],
                    "Semester 2": ["Plant Anatomy", "Plant Physiology", "Botany Lab-2"],
                    "Semester 3": ["Plant Taxonomy", "Plant Ecology", "Botany Lab-3"],
                    "Semester 4": ["Plant Genetics", "Plant Biochemistry", "Botany Lab-4"],
                    "Semester 5": ["Plant Biotechnology", "Plant Pathology", "Botany Lab-5"],
                    "Semester 6": ["Economic Botany", "Plant Conservation", "Botany Lab-6"]
                },
                "Zoology": {
                    "Semester 1": ["Animal Diversity", "Cell Biology", "Zoology Lab-1"],
                    "Semester 2": ["Animal Physiology", "Evolution", "Zoology Lab-2"],
                    "Semester 3": ["Genetics", "Developmental Biology", "Zoology Lab-3"],
                    "Semester 4": ["Ecology", "Animal Behavior", "Zoology Lab-4"],
                    "Semester 5": ["Biotechnology", "Immunology", "Zoology Lab-5"],
                    "Semester 6": ["Applied Zoology", "Wildlife Biology", "Zoology Lab-6"]
                },
                "Microbiology": {
                    "Semester 1": ["Basic Microbiology", "Cell Biology", "Microbiology Lab-1"],
                    "Semester 2": ["Bacteriology", "Virology", "Microbiology Lab-2"],
                    "Semester 3": ["Immunology", "Mycology", "Microbiology Lab-3"],
                    "Semester 4": ["Molecular Biology", "Biochemistry", "Microbiology Lab-4"],
                    "Semester 5": ["Industrial Microbiology", "Medical Microbiology", "Microbiology Lab-5"],
                    "Semester 6": ["Environmental Microbiology", "Food Microbiology", "Microbiology Lab-6"]
                },
                "Statistics": {
                    "Semester 1": ["Descriptive Statistics", "Probability", "Statistical Computing-1"],
                    "Semester 2": ["Statistical Inference", "Sampling Theory", "Statistical Computing-2"],
                    "Semester 3": ["Regression Analysis", "Design of Experiments", "Statistical Computing-3"],
                    "Semester 4": ["Multivariate Analysis", "Time Series", "Statistical Computing-4"],
                    "Semester 5": ["Operations Research", "Biostatistics", "Statistical Computing-5"],
                    "Semester 6": ["Data Analytics", "Statistical Quality Control", "Statistical Computing-6"]
                }
            }

            # Add subjects to database
            for department, semesters in department_subjects.items():
                for semester, subjects in semesters.items():
                    for subject in subjects:
                        syllabus = Syllabus(
                            department=department,
                            semester=semester,
                            subject=subject,
                            units="Unit 1,Unit 2,Unit 3,Unit 4"  # Default units for all subjects
                        )
                        db.session.add(syllabus)
            
            db.session.commit()
            print("Syllabus data initialized successfully!")
    except Exception as e:
        db.session.rollback()
        print(f"Error initializing syllabus data: {str(e)}")

# Initialize database and faculty data
with app.app_context():
    db.create_all()
    init_faculty_data()
    init_syllabus_data()

@app.route('/attendance-history')
@login_required
def attendance_history_page():
    return render_template('attendance_history.html',
                         departments=DEPARTMENTS,
                         subject_codes=SUBJECT_CODES,
                         units=UNITS,
                         semesters=SEMESTERS,
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/get-syllabus-for-attendance/<department>/<semester>')
@login_required
def get_syllabus_for_attendance(department, semester):
    try:
        syllabus_entries = Syllabus.query.filter_by(
            department=department,
            semester=semester
        ).all()
        
        entries = [{
            'id': entry.id,
            'subject': entry.subject,
            'units': entry.units.split(',')
        } for entry in syllabus_entries]
        
        return jsonify({
            'success': True,
            'entries': entries
        })
    except Exception as e:
        print(f"Error fetching syllabus for attendance: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching syllabus: {str(e)}'
        })

@app.route('/get-registered-students')
@admin_login_required
def get_registered_students():
    try:
        students = Student.query.all()
        student_list = []
        
        for student in students:
            photos = []
            if student.photo:
                # Add the main photo
                photos.append(base64.b64encode(student.photo).decode('utf-8'))
                
                # Add the additional photos from face_encoding
                if isinstance(student.face_encoding, list):
                    for encoding in student.face_encoding:
                        if isinstance(encoding, bytes):
                            # If it's bytes, it's a photo
                            photos.append(base64.b64encode(encoding).decode('utf-8'))
                        elif isinstance(encoding, list) and len(encoding) == 4:
                            # If it's a list of 4 numbers, it's face coordinates, skip it
                            continue
                        else:
                            # Try to get the photo data from the encoding
                            try:
                                photo_data = bytes(encoding)
                                photos.append(base64.b64encode(photo_data).decode('utf-8'))
                            except:
                                continue
            
            student_list.append({
                'id': student.id,
                'name': student.name,
                'enrollment_number': student.enrollment_number,
                'semester': student.semester,
                'group': student.group,
                'has_photos': student.photo is not None,
                'photos': photos
            })
        
        return jsonify({
            'success': True,
            'students': student_list
        })
    except Exception as e:
        print(f"Error fetching registered students: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching students: {str(e)}'
        })

@app.route('/update-student-photos', methods=['POST'])
@admin_login_required
def update_student_photos():
    try:
        data = request.json
        if not data or 'student_id' not in data or 'images' not in data:
            return jsonify({'success': False, 'message': 'Missing required data'})

        if len(data['images']) != 3:
            return jsonify({'success': False, 'message': 'Three photos are required'})

        student = db.session.get(Student, data['student_id'])
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'})

        photos = []
        face_encodings = []
        
        for img_data in data['images']:
            img = process_base64_image(img_data)  # Ensure this function handles the image correctly
            
            # Check if the image is valid
            if img is None:
                return jsonify({'success': False, 'message': 'Invalid image format or corrupted image'})
            
            # Detect faces using face_recognition
            face_locations = face_recognition.face_locations(img)
            if len(face_locations) == 0:
                return jsonify({'success': False, 'message': 'No face detected in one or more photos'})
            
            if len(face_locations) > 1:
                return jsonify({'success': False, 'message': 'Multiple faces detected in one photo. Please capture one face at a time'})
            
            # Get face encodings
            face_encoding = face_recognition.face_encodings(img, face_locations)[0]
            
            # Convert image back to BGR for storage
            img_bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            _, img_encoded = cv2.imencode('.jpg', img_bgr)
            photo_bytes = img_encoded.tobytes()
            
            photos.append(photo_bytes)
            face_encodings.append(face_encoding)

        # Update student photos - store all photos and their encodings
        student.photo = photos[0]  # Store the first photo as main photo
        student.face_encoding = face_encodings  # Store face encodings
        
        db.session.commit()

        return jsonify({'success': True, 'message': f'Successfully updated photos for {student.name}'})

    except Exception as e:
        db.session.rollback()
        print(f"Error updating student photos: {str(e)}")
        return jsonify({'success': False, 'message': f'Error updating photos: {str(e)}'})

@app.route('/import-students-excel', methods=['POST'])
@admin_login_required
def import_students_excel():
    try:
        print("\n=== Starting Excel Import Process ===")
        
        if 'file' not in request.files:
            print("No file in request")
            return jsonify({
                'success': False,
                'message': 'No file uploaded'
            })

        file = request.files['file']
        if file.filename == '':
            print("Empty filename")
            return jsonify({
                'success': False,
                'message': 'No file selected'
            })

        if not file.filename.endswith(('.xls', '.xlsx')):
            print("Invalid file type")
            return jsonify({
                'success': False,
                'message': 'Please upload an Excel file (.xls or .xlsx)'
            })

        # Read Excel file
        print(f"Reading Excel file: {file.filename}")
        df = pd.read_excel(file)
        print(f"Excel data columns: {df.columns.tolist()}")
        print(f"Number of rows: {len(df)}")
        
        required_columns = ['Name', 'Enrollment Number', 'Semester', 'Group', 'Gender', 'Category', 'Mobile Number']
        
        # Verify required columns exist
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"Missing columns: {missing_columns}")
            return jsonify({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            })

        # Process each row
        success_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                print(f"\nProcessing row {index + 2}:")
                print(f"Name: {row['Name']}")
                print(f"Enrollment: {row['Enrollment Number']}")
                print(f"Semester: {row['Semester']}")
                print(f"Group: {row['Group']}")
                print(f"Gender: {row['Gender']}")
                print(f"Category: {row['Category']}")
                print(f"Mobile: {row['Mobile Number']}")

                # Data validation
                if not str(row['Name']).strip():
                    raise ValueError("Name is required")
                if not str(row['Enrollment Number']).strip():
                    raise ValueError("Enrollment Number is required")
                if not str(row['Semester']).strip():
                    raise ValueError("Semester is required")
                if not str(row['Group']).strip():
                    raise ValueError("Group is required")
                if not str(row['Gender']).strip():
                    raise ValueError("Gender is required")
                if not str(row['Category']).strip():
                    raise ValueError("Category is required")
                if not str(row['Mobile Number']).strip():
                    raise ValueError("Mobile Number is required")

                # Check if student already exists
                existing_student = Student.query.filter_by(
                    enrollment_number=str(row['Enrollment Number']).strip()
                ).first()

                if existing_student:
                    error_msg = f"Student with enrollment {row['Enrollment Number']} already exists"
                    print(error_msg)
                    errors.append(error_msg)
                    continue

                # Validate group and get subject mappings
                group = str(row['Group']).strip()
                if group not in GROUP_SUBJECTS:
                    error_msg = f"Invalid group {group} for student {row['Enrollment Number']}"
                    print(error_msg)
                    errors.append(error_msg)
                    continue

                subjects = GROUP_SUBJECTS[group]
                print(f"Subject mappings: {subjects}")

                # Create new student
                new_student = Student(
                    name=str(row['Name']).strip(),
                    enrollment_number=str(row['Enrollment Number']).strip(),
                    semester=str(row['Semester']).strip(),
                    group=group,
                    major_subject=subjects['major'],
                    minor_subject=subjects['minor'],
                    multi_subject=subjects['multi'],
                    gender=str(row['Gender']).strip(),
                    category=str(row['Category']).strip(),
                    mobile_number=str(row['Mobile Number']).strip()
                )
                
                db.session.add(new_student)
                success_count += 1
                print(f"Successfully added student: {new_student.name}")

            except Exception as e:
                error_msg = f"Error processing row {index + 2}: {str(e)}"
                print(error_msg)
                errors.append(error_msg)
                continue

        db.session.commit()
        print(f"\nImport completed. Success count: {success_count}, Errors: {len(errors)}")

        return jsonify({
            'success': True,
            'message': f'Successfully imported {success_count} students',
            'errors': errors if errors else None
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error importing students: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error importing students: {str(e)}'
        })

# HOD Timetable Management Routes
@app.route('/manage-timetable')
@login_required
def manage_timetable():
    if not session.get('is_hod'):
        return redirect(url_for('index'))
    
    department = session.get('department')
    try:
        # Get all faculty members in the department
        faculty_members = Faculty.query.filter_by(department=department).all()
        
        # Get department syllabus for all semesters
        syllabus = {}
        for semester in SEMESTERS:
            syllabus_entries = Syllabus.query.filter_by(
                department=department,
                semester=semester
            ).all()
            syllabus[semester] = [{
                'id': entry.id,
                'subject': entry.subject,
                'units': entry.units.split(',') if entry.units else []
            } for entry in syllabus_entries]
        
        # Get existing timetable
        timetable = Timetable.query.filter_by(department=department).all()
        
        print(f"Loaded syllabus for {department}: {syllabus}")  # Debug print
        
        return render_template('manage_timetable.html',
                             faculty_members=faculty_members,
                             syllabus=syllabus,
                             timetable=timetable,
                             semesters=SEMESTERS,
                             department=department)
    except Exception as e:
        print(f"Error loading timetable management: {str(e)}")
        return redirect(url_for('index'))

@app.route('/add-timetable-entry', methods=['POST'])
@login_required
def add_timetable_entry():
    if not session.get('is_hod'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        data = request.json
        
        # Validate time slot format
        if not data.get('time_slot') or not data.get('day'):
            return jsonify({
                'success': False,
                'message': 'Time slot and day are required'
            })

        # Check for conflicts
        existing = Timetable.query.filter_by(
            department=session.get('department'),
            semester=data['semester'],
            day=data['day'],
            time_slot=data['time_slot']
        ).first()
        
        if existing:
            return jsonify({
                'success': False,
                'message': 'Time slot already occupied'
            })

        # Create new entry
        new_entry = Timetable(
            department=session.get('department'),
            semester=data['semester'],
            day=data['day'],
            time_slot=data['time_slot'],
            subject=data['subject'],
            faculty_id=data['faculty_id'],
            unit=data['unit']
        )
        
        db.session.add(new_entry)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Timetable entry added successfully'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error adding timetable entry: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error adding entry: {str(e)}'
        })

@app.route('/get-timetable/<semester>')
@login_required
def get_timetable(semester):
    try:
        department = session.get('department')
        timetable = Timetable.query.filter_by(
            department=department,
            semester=semester
        ).all()
        
        entries = []
        for entry in timetable:
            faculty = Faculty.query.get(entry.faculty_id)
            entries.append({
                'id': entry.id,
                'day': entry.day,
                'time_slot': entry.time_slot,
                'subject': entry.subject,
                'faculty_name': faculty.name if faculty else 'Unknown',
                'unit': entry.unit
            })
        
        return jsonify({
            'success': True,
            'timetable': entries
        })
    except Exception as e:
        print(f"Error fetching timetable: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching timetable: {str(e)}'
        })

@app.route('/delete-timetable-entry/<int:entry_id>', methods=['POST'])
@login_required
def delete_timetable_entry(entry_id):
    if not session.get('is_hod'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        entry = Timetable.query.get_or_404(entry_id)
        if entry.department != session.get('department'):
            return jsonify({'success': False, 'message': 'Unauthorized'})
        
        db.session.delete(entry)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Timetable entry deleted successfully'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting timetable entry: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error deleting entry: {str(e)}'
        })

@app.route('/get-subjects/<semester>')
def get_subjects(semester):
    try:
        department = session.get('department')
        subjects = Syllabus.query.filter_by(
            department=department,
            semester=semester
        ).all()
        
        subject_list = [{
            'id': subject.id,
            'subject': subject.subject,
            'units': subject.units.split(',') if subject.units else []
        } for subject in subjects]
        
        print(f"Fetched subjects for {department}, {semester}: {subject_list}")  # Debug print
        
        return jsonify({
            'success': True,
            'subjects': subject_list
        })
    except Exception as e:
        print(f"Error fetching subjects: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching subjects: {str(e)}'
        })

@app.route('/get-faculty-schedule')
@login_required
def get_faculty_schedule():
    try:
        faculty_id = session.get('faculty_id')
        department = session.get('department')
        
        # Get all timetable entries for this faculty
        schedule = Timetable.query.filter_by(
            department=department,
            faculty_id=faculty_id
        ).all()
        
        schedule_list = [{
            'day': entry.day,
            'time_slot': entry.time_slot,
            'semester': entry.semester,
            'subject': entry.subject,
            'unit': entry.unit
        } for entry in schedule]
        
        return jsonify({
            'success': True,
            'schedule': schedule_list
        })
    except Exception as e:
        print(f"Error fetching faculty schedule: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching schedule: {str(e)}'
        })

@app.route('/get-all-students')
@login_required
def get_all_students():
    try:
        students = Student.query.all()
        student_list = []
        
        for student in students:
            student_list.append({
                'id': student.id,
                'name': student.name,
                'enrollment_number': student.enrollment_number,
                'semester': student.semester,
                'group': student.group,
                'major_subject': student.major_subject,
                'minor_subject': student.minor_subject,
                'multi_subject': student.multi_subject,
                'gender': student.gender,
                'category': student.category,
                'mobile_number': student.mobile_number
            })
        
        return jsonify({
            'success': True,
            'students': student_list
        })
    except Exception as e:
        print(f"Error fetching students: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching students: {str(e)}'
        })

@app.route('/get-all-subjects')
@login_required
def get_all_subjects():
    try:
        department = session.get('department')
        subjects = Syllabus.query.filter_by(department=department).all()
        
        subject_list = [{
            'id': subject.id,
            'subject': subject.subject,
            'semester': subject.semester
        } for subject in subjects]
        
        return jsonify({
            'success': True,
            'subjects': subject_list
        })
    except Exception as e:
        print(f"Error fetching all subjects: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching subjects: {str(e)}'
        })

@app.route('/get-subject-details/<int:subject_id>')
@login_required
def get_subject_details(subject_id):
    try:
        subject = Syllabus.query.get_or_404(subject_id)
        
        # Check if subject belongs to faculty's department
        if subject.department != session.get('department'):
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            })
        
        return jsonify({
            'success': True,
            'subject': {
                'id': subject.id,
                'subject': subject.subject,
                'semester': subject.semester,
                'units': subject.units.split(',') if subject.units else []
            }
        })
    except Exception as e:
        print(f"Error fetching subject details: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching subject details: {str(e)}'
        })

@app.route('/complete-unit', methods=['POST'])
@login_required
def complete_unit():
    try:
        data = request.json
        subject_id = data.get('subject_id')
        unit = data.get('unit')
        
        if not subject_id or not unit:
            return jsonify({
                'success': False,
                'message': 'Subject ID and unit are required'
            })
        
        subject = Syllabus.query.get_or_404(subject_id)
        
        # Check if subject belongs to faculty's department
        if subject.department != session.get('department'):
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            })
        
        # Create or update completion record
        completion = SubjectCompletion.query.filter_by(
            subject_id=subject_id,
            unit=unit,
            faculty_id=session.get('faculty_id')
        ).first()
        
        if not completion:
            completion = SubjectCompletion(
                subject_id=subject_id,
                unit=unit,
                faculty_id=session.get('faculty_id'),
                completed_at=datetime.now()
            )
            db.session.add(completion)
        else:
            completion.completed_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Unit marked as complete'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error completing unit: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error completing unit: {str(e)}'
        })

@app.route('/complete-subject', methods=['POST'])
@login_required
def complete_subject():
    try:
        data = request.json
        subject_id = data.get('subject_id')
        
        if not subject_id:
            return jsonify({
                'success': False,
                'message': 'Subject ID is required'
            })
        
        subject = Syllabus.query.get_or_404(subject_id)
        
        # Check if subject belongs to faculty's department
        if subject.department != session.get('department'):
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            })
        
        # Get all units for the subject
        units = subject.units.split(',') if subject.units else []
        
        # Create or update completion records for all units
        for unit in units:
            completion = SubjectCompletion.query.filter_by(
                subject_id=subject_id,
                unit=unit,
                faculty_id=session.get('faculty_id')
            ).first()
            
            if not completion:
                completion = SubjectCompletion(
                    subject_id=subject_id,
                    unit=unit,
                    faculty_id=session.get('faculty_id'),
                    completed_at=datetime.now()
                )
                db.session.add(completion)
            else:
                completion.completed_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Subject marked as complete'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error completing subject: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error completing subject: {str(e)}'
        })

def capture_image():
    try:
        # Force DirectShow backend
        cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        if not cap.isOpened():
            print("Error: Could not open webcam with DirectShow backend")
            return None

        # Set specific properties for DirectShow
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        cap.set(cv2.CAP_PROP_FPS, 30)
        cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'))
        
        # Give the camera time to warm up
        time.sleep(1.0)
        
        # Try to capture frame
        ret, frame = cap.read()
        if not ret or frame is None:
            print("Error: Could not read frame from webcam")
            cap.release()
            return None

        # Convert BGR to RGB
        image_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        print("Successfully captured image from webcam")
        
        # Release the camera
        cap.release()
        return image_rgb

    except Exception as e:
        print(f"Error in capture_image: {str(e)}")
        if 'cap' in locals() and cap is not None:
            cap.release()
        return None

@app.route('/backup-database', methods=['POST'])
@admin_login_required
def backup_database():
    try:
        print("Starting database backup process...")
        
        # Create backups directory if it doesn't exist
        current_dir = os.path.dirname(os.path.abspath(__file__))
        backups_dir = os.path.join(current_dir, 'backups')
        print(f"Creating backups directory at: {backups_dir}")
        os.makedirs(backups_dir, exist_ok=True)

        # Create timestamp for backup files
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        print(f"Backup timestamp: {timestamp}")
        
        # Create Excel writer
        excel_filename = f'database_backup_{timestamp}.xlsx'
        excel_path = os.path.join(backups_dir, excel_filename)
        print(f"Creating Excel backup at: {excel_path}")
        writer = pd.ExcelWriter(excel_path, engine='xlsxwriter')
        
        # Backup Students table
        print("Backing up students data...")
        students = Student.query.all()
        print(f"Found {len(students)} students in database")
        
        student_data = []
        for student in students:
            try:
                # Create student dictionary with all relevant fields
                student_dict = {
                    'id': student.id,
                    'name': student.name,
                    'enrollment_number': student.enrollment_number,
                    'semester': student.semester,
                    'group': student.group,
                    'major_subject': student.major_subject,
                    'minor_subject': student.minor_subject,
                    'multi_subject': student.multi_subject,
                    'gender': student.gender,
                    'category': student.category,
                    'mobile_number': student.mobile_number,
                    'has_photo': student.photo is not None,
                    'has_face_encoding': student.face_encoding is not None
                }
                student_data.append(student_dict)
                print(f"Processed student: {student.name} (ID: {student.id})")
            except Exception as e:
                print(f"Error processing student {student.id}: {str(e)}")
                continue
        
        if student_data:
            print(f"Creating Students sheet with {len(student_data)} records")
            pd.DataFrame(student_data).to_excel(writer, sheet_name='Students', index=False)
        else:
            print("No student data to backup")
        
        # Backup Attendance records
        print("Backing up attendance data...")
        attendance_data = (
            db.session.query(
                Attendance.id,
                Attendance.student_id,
                Attendance.date,
                Attendance.time,
                Attendance.department,
                Attendance.faculty,
                Attendance.subject,
                Attendance.unit
            ).all()
        )
        attendance_df = pd.DataFrame(attendance_data, columns=[
            'id', 'student_id', 'date', 'time', 'department', 'faculty', 'subject', 'unit'
        ])
        attendance_df.to_excel(writer, sheet_name='Attendance', index=False)
        print(f"Backed up {len(attendance_data)} attendance records")
        
        # Backup Faculty data
        print("Backing up faculty data...")
        faculty_data = Faculty.query.all()
        faculty_list = [{
            'id': f.id,
            'name': f.name,
            'department': f.department,
            'is_hod': f.is_hod
        } for f in faculty_data]
        pd.DataFrame(faculty_list).to_excel(writer, sheet_name='Faculty', index=False)
        print(f"Backed up {len(faculty_data)} faculty members")
        
        # Backup Syllabus data
        print("Backing up syllabus data...")
        syllabus_data = Syllabus.query.all()
        syllabus_list = [{
            'id': s.id,
            'department': s.department,
            'semester': s.semester,
            'subject': s.subject,
            'units': s.units
        } for s in syllabus_data]
        pd.DataFrame(syllabus_list).to_excel(writer, sheet_name='Syllabus', index=False)
        print(f"Backed up {len(syllabus_data)} syllabus entries")
        
        # Save Excel file
        print("Saving Excel file...")
        writer.close()
        
        # Create a ZIP file containing the Excel backup
        zip_filename = f'full_backup_{timestamp}.zip'
        zip_path = os.path.join(backups_dir, zip_filename)
        print(f"Creating ZIP file at: {zip_path}")
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(excel_path, os.path.basename(excel_path))
            print("Added Excel file to ZIP")
            
            # Add SQLite database file to backup if using SQLite
            db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
            if os.path.exists(db_path):
                zipf.write(db_path, f'database_{timestamp}.db')
                print("Added database file to ZIP")
        
        # Read the ZIP file
        print("Reading ZIP file for response...")
        with open(zip_path, 'rb') as f:
            zip_data = f.read()
        
        # Create response with proper headers
        print("Creating response...")
        response = make_response(zip_data)
        response.headers['Content-Type'] = 'application/zip'
        response.headers['Content-Disposition'] = f'attachment; filename={zip_filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        # Add Excel file path to response headers
        response.headers['X-Excel-File'] = excel_filename
        
        print("Backup process completed successfully")
        return response
        
    except Exception as e:
        print(f"Error in backup process: {str(e)}")
        print(f"Error type: {type(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'Error creating backup: {str(e)}'
        }), 500

@app.route('/restore-database', methods=['POST'])
@admin_login_required
def restore_database():
    try:
        if 'backup_file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No backup file provided'
            })
            
        backup_file = request.files['backup_file']
        if not backup_file.filename.endswith('.zip'):
            return jsonify({
                'success': False,
                'message': 'Invalid backup file format. Please provide a ZIP file.'
            })
            
        # Create temporary directory for extraction
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Extract ZIP file
            with zipfile.ZipFile(backup_file, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Find Excel file in extracted contents
            excel_files = glob.glob(os.path.join(temp_dir, '*.xlsx'))
            if not excel_files:
                raise ValueError('No Excel backup file found in ZIP')
                
            excel_path = excel_files[0]
            
            # Begin database restoration
            db.session.query(Student).delete()
            db.session.query(Attendance).delete()
            db.session.query(Faculty).delete()
            db.session.query(Syllabus).delete()
            
            # Restore Students
            students_df = pd.read_excel(excel_path, sheet_name='Students')
            for _, row in students_df.iterrows():
                photo_data = base64.b64decode(row['photo_data']) if pd.notna(row['photo_data']) else None
                face_encoding = json.loads(row['face_encoding']) if pd.notna(row['face_encoding']) else None
                
                student = Student(
                    name=row['name'],
                    enrollment_number=row['enrollment_number'],
                    department=row['department'],
                    semester=row['semester'],
                    photo=photo_data,
                    face_encoding=face_encoding
                )
                db.session.add(student)
            
            # Restore Attendance
            attendance_df = pd.read_excel(excel_path, sheet_name='Attendance')
            for _, row in attendance_df.iterrows():
                attendance = Attendance(
                    student_id=row['student_id'],
                    date=row['date'],
                    time=row['time'],
                    department=row['department'],
                    faculty=row['faculty'],
                    subject=row['subject'],
                    unit=row['unit']
                )
                db.session.add(attendance)
            
            # Restore Faculty
            faculty_df = pd.read_excel(excel_path, sheet_name='Faculty')
            for _, row in faculty_df.iterrows():
                faculty = Faculty(
                    name=row['name'],
                    department=row['department'],
                    is_hod=row['is_hod']
                )
                db.session.add(faculty)
            
            # Restore Syllabus
            syllabus_df = pd.read_excel(excel_path, sheet_name='Syllabus')
            for _, row in syllabus_df.iterrows():
                syllabus = Syllabus(
                    department=row['department'],
                    semester=row['semester'],
                    subject=row['subject'],
                    units=json.loads(row['units'])
                )
                db.session.add(syllabus)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Database restored successfully'
            })
            
        finally:
            # Clean up temporary directory
            shutil.rmtree(temp_dir)
            
    except Exception as e:
        db.session.rollback()
        print(f"Restore error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error restoring database: {str(e)}'
        })

@app.route('/backups/<path:filename>')
@admin_login_required
def serve_backup(filename):
    try:
        backups_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
        return send_from_directory(backups_dir, filename)
    except Exception as e:
        print(f"Error serving backup file: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error serving backup file: {str(e)}'
        }), 404

if __name__ == "__main__":
    app.run(debug=True, threaded=False)

